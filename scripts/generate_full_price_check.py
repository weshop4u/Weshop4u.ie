import json
import subprocess
import urllib.parse
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Priority order for category tabs
CATEGORY_PRIORITY_ORDER = [
    "Deli",
    "Fizzy Drinks",
    "Energy Drinks",
    "Water and Flavoured Water",
    "Chocolate Bars",
    "Chocolates Multi packs and Boxes",
    "Crisps and Nuts",
    "Biscuits and Cookies",
    "Tobacco and Cigars and Papers",
    "Vapes and Vape Oils",
    "Spirits",
    "Cans and Bottles",
    "Flavored Alcohol",
    "Wines",
    "Nicotine Products",
]

# Fetch products from API
input_json = json.dumps({"json": {"storeId": 1}})
encoded = urllib.parse.quote(input_json)
url = f"http://127.0.0.1:3000/api/trpc/store.getProducts?input={encoded}"
result = subprocess.run(["curl", "-s", url], capture_output=True, text=True)
data = json.loads(result.stdout)
products = data["result"]["data"]["json"]

# Only active products
active_products = [p for p in products if p.get("isActive", True)]

# Group by category
categories = {}
for p in active_products:
    cat = p.get("categoryName") or "Uncategorized"
    if cat not in categories:
        categories[cat] = []
    categories[cat].append(p)

# Sort categories: priority first, then alphabetical
def cat_sort_key(cat_name):
    name_lower = cat_name.lower().strip()
    for i, priority in enumerate(CATEGORY_PRIORITY_ORDER):
        if name_lower == priority.lower().strip():
            return (0, i, cat_name)
    return (1, 0, cat_name.lower())

sorted_cats = sorted(categories.keys(), key=cat_sort_key)

# Skip Fizzy Drinks (already done)
sorted_cats = [c for c in sorted_cats if c != "Fizzy Drinks"]

print(f"Total categories to generate: {len(sorted_cats)}")
print(f"Total products (excl Fizzy Drinks): {sum(len(categories[c]) for c in sorted_cats)}")

# Styles
header_font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
header_fill = PatternFill(start_color="1B7A8A", end_color="1B7A8A", fill_type="solid")
title_font = Font(name="Calibri", bold=True, size=12, color="1B7A8A")
subtitle_font = Font(name="Calibri", size=8, color="666666")
data_font = Font(name="Calibri", size=9)
price_font = Font(name="Calibri", size=9)
num_font = Font(name="Calibri", size=8, color="888888")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
alt_fill = PatternFill(start_color="F0F6F7", end_color="F0F6F7", fill_type="solid")

# New Products sheet styles
new_header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
new_title_font = Font(name="Calibri", bold=True, size=12, color="2E7D32")

wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

def sanitize_sheet_name(name):
    """Excel sheet names max 31 chars, no special chars"""
    # Replace problematic chars
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, '')
    return name[:31]

def create_category_sheet(wb, cat_name, cat_products):
    """Create a two-column price check sheet for a category."""
    sheet_name = sanitize_sheet_name(cat_name)
    ws = wb.create_sheet(title=sheet_name)
    
    # Sort products alphabetically
    cat_products = sorted(cat_products, key=lambda x: x["name"].lower())
    
    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 1.5
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 32
    ws.column_dimensions["H"].width = 9
    ws.column_dimensions["I"].width = 9

    # Title row
    ws.merge_cells("A1:I1")
    cell = ws["A1"]
    cell.value = f"SPAR BALBRIGGAN — {cat_name}"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Subtitle row
    ws.merge_cells("A2:I2")
    cell = ws["A2"]
    cell.value = f"{len(cat_products)} products · Price Check Sheet · Date: ___/___/2026 · Checked by: _______________"
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Header row (row 3)
    HEADER_ROW = 3
    left_cols = [1, 2, 3, 4]
    right_cols = [6, 7, 8, 9]
    headers = ["#", "Product Name", "App €", "Shelf €"]

    for col, header in zip(left_cols, headers):
        cell = ws.cell(row=HEADER_ROW, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
        cell.border = thin_border

    for col, header in zip(right_cols, headers):
        cell = ws.cell(row=HEADER_ROW, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col != 7 else "left", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[HEADER_ROW].height = 18

    # Split products into left and right halves
    half = math.ceil(len(cat_products) / 2)
    left_products = cat_products[:half]
    right_products = cat_products[half:]

    DATA_START = HEADER_ROW + 1

    def write_product_row(row, col_offset, product, index, use_alt):
        base = col_offset + 1
        fill = alt_fill if use_alt else None

        cell = ws.cell(row=row, column=base, value=index)
        cell.font = num_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if fill: cell.fill = fill

        cell = ws.cell(row=row, column=base + 1, value=product["name"])
        cell.font = data_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        if fill: cell.fill = fill

        price = product.get("price")
        cell = ws.cell(row=row, column=base + 2, value=float(price) if price else "")
        cell.font = price_font
        cell.number_format = '€#,##0.00'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if fill: cell.fill = fill

        cell = ws.cell(row=row, column=base + 3, value="")
        cell.font = data_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if fill: cell.fill = fill

    for i, product in enumerate(left_products):
        row = DATA_START + i
        ws.row_dimensions[row].height = 16
        write_product_row(row, 0, product, i + 1, i % 2 == 1)

    for i, product in enumerate(right_products):
        row = DATA_START + i
        write_product_row(row, 5, product, half + i + 1, i % 2 == 1)

    max_rows = max(len(left_products), len(right_products))
    summary_row = DATA_START + max_rows + 1
    ws.merge_cells(f"A{summary_row}:I{summary_row}")
    cell = ws.cell(row=summary_row, column=1,
        value=f"Total: {len(cat_products)} products  |  Price changes needed: ___  |  Missing from app: ___  |  Remove from app: ___")
    cell.font = Font(name="Calibri", size=8, italic=True, color="666666")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[summary_row].height = 18

    # Print settings
    ws.print_area = f"A1:I{summary_row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"1:{HEADER_ROW}"

    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15

    return len(cat_products)

# Create sheets for each category
total_products = 0
for cat_name in sorted_cats:
    count = create_category_sheet(wb, cat_name, categories[cat_name])
    total_products += count
    print(f"  {cat_name}: {count} products")

# Create "New Products" sheet at the end
ws_new = wb.create_sheet(title="New Products")
ws_new.column_dimensions["A"].width = 4
ws_new.column_dimensions["B"].width = 35
ws_new.column_dimensions["C"].width = 25
ws_new.column_dimensions["D"].width = 10
ws_new.column_dimensions["E"].width = 35
ws_new.column_dimensions["F"].width = 10

ws_new.merge_cells("A1:F1")
cell = ws_new["A1"]
cell.value = "SPAR BALBRIGGAN — New Products (Not in App)"
cell.font = new_title_font
cell.alignment = Alignment(horizontal="left", vertical="center")
ws_new.row_dimensions[1].height = 22

ws_new.merge_cells("A2:F2")
cell = ws_new["A2"]
cell.value = "Products found on shelf but missing from the app · Date: ___/___/2026 · Checked by: _______________"
cell.font = subtitle_font
cell.alignment = Alignment(horizontal="left", vertical="center")
ws_new.row_dimensions[2].height = 16

new_headers = ["#", "Product Name", "Category", "Price €", "Description / Notes", "DRS?"]
for col, header in enumerate(new_headers, 1):
    cell = ws_new.cell(row=3, column=col, value=header)
    cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    cell.alignment = Alignment(horizontal="center" if col != 2 and col != 5 else "left", vertical="center")
    cell.border = thin_border
ws_new.row_dimensions[3].height = 18

# 60 blank rows for writing new products
for i in range(1, 61):
    row = i + 3
    ws_new.row_dimensions[row].height = 18
    cell = ws_new.cell(row=row, column=1, value=i)
    cell.font = num_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if i % 2 == 0:
        cell.fill = alt_fill
    for col in range(2, 7):
        cell = ws_new.cell(row=row, column=col, value="")
        cell.border = thin_border
        if i % 2 == 0:
            cell.fill = alt_fill

ws_new.print_area = "A1:F63"
ws_new.page_setup.orientation = "portrait"
ws_new.page_setup.paperSize = ws_new.PAPERSIZE_A4
ws_new.page_setup.fitToWidth = 1
ws_new.page_setup.fitToHeight = 0
ws_new.sheet_properties.pageSetUpPr.fitToPage = True
ws_new.print_title_rows = "1:3"
ws_new.page_margins.left = 0.3
ws_new.page_margins.right = 0.3
ws_new.page_margins.top = 0.3
ws_new.page_margins.bottom = 0.3

output_path = "/home/ubuntu/weshop4u/Spar_Price_Check_All_Categories.xlsx"
wb.save(output_path)
print(f"\nSaved to {output_path}")
print(f"Total sheets: {len(sorted_cats) + 1} ({len(sorted_cats)} categories + New Products)")
print(f"Total products: {total_products}")
print(f"Estimated paper (double-sided, ~90 per side): ~{math.ceil(total_products / 180)} sheets")
