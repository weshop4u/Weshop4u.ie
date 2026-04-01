import json
import subprocess
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Fetch products from API
input_json = json.dumps({"json": {"storeId": 1}})
encoded = urllib.parse.quote(input_json)
url = f"http://127.0.0.1:3000/api/trpc/store.getProducts?input={encoded}"
result = subprocess.run(["curl", "-s", url], capture_output=True, text=True)
data = json.loads(result.stdout)
products = data["result"]["data"]["json"]

# Filter Fizzy Drinks, sort alphabetically
fizzy = sorted(
    [p for p in products if p.get("categoryName") == "Fizzy Drinks"],
    key=lambda x: x["name"].lower()
)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Fizzy Drinks"

# Styles
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1B7A8A", end_color="1B7A8A", fill_type="solid")
title_font = Font(name="Calibri", bold=True, size=14, color="1B7A8A")
subtitle_font = Font(name="Calibri", size=10, color="666666")
data_font = Font(name="Calibri", size=10)
price_font = Font(name="Calibri", size=10, bold=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
alt_fill = PatternFill(start_color="F5F9FA", end_color="F5F9FA", fill_type="solid")

# Column widths
ws.column_dimensions["A"].width = 5    # #
ws.column_dimensions["B"].width = 42   # Product Name
ws.column_dimensions["C"].width = 14   # App Price
ws.column_dimensions["D"].width = 14   # Shelf Price
ws.column_dimensions["E"].width = 12   # Correct?
ws.column_dimensions["F"].width = 20   # Notes

# Title row
ws.merge_cells("A1:F1")
cell = ws["A1"]
cell.value = "SPAR BALBRIGGAN — Fizzy Drinks"
cell.font = title_font
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

# Subtitle row
ws.merge_cells("A2:F2")
cell = ws["A2"]
cell.value = f"{len(fizzy)} products · Price Check Sheet · Date: ___/___/2026 · Checked by: _______________"
cell.font = subtitle_font
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 22

# Empty row
ws.row_dimensions[3].height = 8

# Header row
headers = ["#", "Product Name", "App Price (€)", "Shelf Price (€)", "Correct?", "Notes"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
    cell.border = thin_border
ws.row_dimensions[4].height = 24

# Data rows
for i, product in enumerate(fizzy):
    row = i + 5
    ws.row_dimensions[row].height = 20
    
    use_fill = alt_fill if i % 2 == 1 else None
    
    # #
    cell = ws.cell(row=row, column=1, value=i + 1)
    cell.font = data_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if use_fill:
        cell.fill = use_fill
    
    # Product Name
    cell = ws.cell(row=row, column=2, value=product["name"])
    cell.font = data_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border
    if use_fill:
        cell.fill = use_fill
    
    # App Price
    price = product.get("price")
    cell = ws.cell(row=row, column=3, value=float(price) if price else "")
    cell.font = price_font
    cell.number_format = '€#,##0.00'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if use_fill:
        cell.fill = use_fill
    
    # Shelf Price (blank)
    cell = ws.cell(row=row, column=4, value="")
    cell.font = data_font
    cell.number_format = '€#,##0.00'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if use_fill:
        cell.fill = use_fill
    
    # Correct? (blank)
    cell = ws.cell(row=row, column=5, value="")
    cell.font = data_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if use_fill:
        cell.fill = use_fill
    
    # Notes (blank)
    cell = ws.cell(row=row, column=6, value="")
    cell.font = data_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border
    if use_fill:
        cell.fill = use_fill

# Summary row at bottom
summary_row = len(fizzy) + 6
ws.merge_cells(f"A{summary_row}:F{summary_row}")
cell = ws.cell(row=summary_row, column=1, value=f"Total: {len(fizzy)} products  |  Wrong prices: ___  |  Missing products: ___  |  Products to remove: ___")
cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[summary_row].height = 24

# Print settings
ws.print_area = f"A1:F{summary_row}"
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = "1:4"  # Repeat header on every page

# Margins (narrow)
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.4
ws.page_margins.header = 0.2
ws.page_margins.footer = 0.2

output_path = "/home/ubuntu/weshop4u/Spar_Fizzy_Drinks_Price_Check.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Products: {len(fizzy)}")
print(f"Estimated pages: ~{max(1, (len(fizzy) + 34) // 35)}")
