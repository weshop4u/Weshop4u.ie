import json
import subprocess
import urllib.parse
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Fetch products from API
input_json = json.dumps({"json": {"storeId": 1}})
encoded = urllib.parse.quote(input_json)
url = f"http://127.0.0.1:3000/api/trpc/store.getProducts?input={encoded}"
result = subprocess.run(["curl", "-s", url], capture_output=True, text=True)
data = json.loads(result.stdout)
products = data["result"]["data"]["json"]

# Filter Fizzy Drinks, sort alphabetically
fizzy = sorted(
    [p for p in products if p.get("categoryName") == "Fizzy Drinks"],
    key=lambda x: x["name"].lower()
)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Fizzy Drinks"

# Styles
header_font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
header_fill = PatternFill(start_color="1B7A8A", end_color="1B7A8A", fill_type="solid")
title_font = Font(name="Calibri", bold=True, size=12, color="1B7A8A")
subtitle_font = Font(name="Calibri", size=8, color="666666")
data_font = Font(name="Calibri", size=9)
price_font = Font(name="Calibri", size=9)
num_font = Font(name="Calibri", size=8, color="888888")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
alt_fill = PatternFill(start_color="F0F6F7", end_color="F0F6F7", fill_type="solid")

# Column layout: A-D = left table, E = spacer, F-I = right table
# A: # (left)    B: Product Name (left)    C: App € (left)    D: Shelf € (left)
# E: spacer
# F: # (right)   G: Product Name (right)   H: App € (right)   I: Shelf € (right)

ws.column_dimensions["A"].width = 4      # #
ws.column_dimensions["B"].width = 32     # Product Name
ws.column_dimensions["C"].width = 9      # App Price
ws.column_dimensions["D"].width = 9      # Shelf Price
ws.column_dimensions["E"].width = 1.5    # Spacer
ws.column_dimensions["F"].width = 4      # #
ws.column_dimensions["G"].width = 32     # Product Name
ws.column_dimensions["H"].width = 9      # App Price
ws.column_dimensions["I"].width = 9      # Shelf Price

# Title row
ws.merge_cells("A1:I1")
cell = ws["A1"]
cell.value = "SPAR BALBRIGGAN — Fizzy Drinks"
cell.font = title_font
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 22

# Subtitle row
ws.merge_cells("A2:I2")
cell = ws["A2"]
cell.value = f"{len(fizzy)} products · Price Check Sheet · Date: ___/___/2026 · Checked by: _______________"
cell.font = subtitle_font
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 16

# Header row (row 3)
HEADER_ROW = 3
left_headers = ["#", "Product Name", "App €", "Shelf €"]
right_headers = ["#", "Product Name", "App €", "Shelf €"]
left_cols = [1, 2, 3, 4]   # A, B, C, D
right_cols = [6, 7, 8, 9]  # F, G, H, I

for col, header in zip(left_cols, left_headers):
    cell = ws.cell(row=HEADER_ROW, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
    cell.border = thin_border

for col, header in zip(right_cols, right_headers):
    cell = ws.cell(row=HEADER_ROW, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center" if col != 7 else "left", vertical="center")
    cell.border = thin_border

ws.row_dimensions[HEADER_ROW].height = 18

# Split products into left and right halves
ROWS_PER_PAGE_SIDE = 33  # rows of data per page side (accounting for title + header)
half = math.ceil(len(fizzy) / 2)
left_products = fizzy[:half]
right_products = fizzy[half:]

DATA_START = HEADER_ROW + 1

def write_product_row(row, col_offset, product, index, use_alt):
    """Write a single product into a row at the given column offset."""
    # col_offset: 0 for left (cols A-D), 5 for right (cols F-I)
    base = col_offset + 1
    fill = alt_fill if use_alt else None

    # #
    cell = ws.cell(row=row, column=base, value=index)
    cell.font = num_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if fill:
        cell.fill = fill

    # Product Name
    cell = ws.cell(row=row, column=base + 1, value=product["name"])
    cell.font = data_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border
    if fill:
        cell.fill = fill

    # App Price
    price = product.get("price")
    cell = ws.cell(row=row, column=base + 2, value=float(price) if price else "")
    cell.font = price_font
    cell.number_format = '€#,##0.00'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if fill:
        cell.fill = fill

    # Shelf Price (blank for writing)
    cell = ws.cell(row=row, column=base + 3, value="")
    cell.font = data_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if fill:
        cell.fill = fill

# Write left side products
for i, product in enumerate(left_products):
    row = DATA_START + i
    ws.row_dimensions[row].height = 16
    write_product_row(row, 0, product, i + 1, i % 2 == 1)

# Write right side products
for i, product in enumerate(right_products):
    row = DATA_START + i
    write_product_row(row, 5, product, half + i + 1, i % 2 == 1)

# Summary row at bottom
max_rows = max(len(left_products), len(right_products))
summary_row = DATA_START + max_rows + 1
ws.merge_cells(f"A{summary_row}:I{summary_row}")
cell = ws.cell(row=summary_row, column=1,
    value=f"Total: {len(fizzy)} products  |  Price changes needed: ___  |  Missing from app: ___  |  Remove from app: ___")
cell.font = Font(name="Calibri", size=8, italic=True, color="666666")
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[summary_row].height = 18

# Print settings
ws.print_area = f"A1:I{summary_row}"
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = f"1:{HEADER_ROW}"  # Repeat title + header on every page

# Narrow margins
ws.page_margins.left = 0.3
ws.page_margins.right = 0.3
ws.page_margins.top = 0.3
ws.page_margins.bottom = 0.3
ws.page_margins.header = 0.15
ws.page_margins.footer = 0.15

output_path = "/home/ubuntu/weshop4u/Spar_Fizzy_Drinks_Price_Check_v2.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Products: {len(fizzy)}")
print(f"Left side: {len(left_products)} products (1-{len(left_products)})")
print(f"Right side: {len(right_products)} products ({half+1}-{len(fizzy)})")
print(f"Data rows needed: {max_rows}")
pages_estimate = math.ceil(max_rows / ROWS_PER_PAGE_SIDE)
print(f"Estimated page sides: ~{pages_estimate} (so ~{math.ceil(pages_estimate/2)} sheets front+back)")
